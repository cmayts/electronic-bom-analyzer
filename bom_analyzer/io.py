"""Input readers and report writers."""

from __future__ import annotations

import csv
import html
import json
from pathlib import Path
from typing import Any


def read_bom(path: Path) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as stream:
            return list(csv.DictReader(stream))
    if suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("XLSX support requires: pip install openpyxl") from exc
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(rows, ())]
        return [dict(zip(headers, row, strict=False)) for row in rows]
    raise ValueError(f"Unsupported input format: {suffix or '<none>'}")


def write_json(path: Path, payload: dict[str, Any]) -> None:
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


def write_normalized_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    fields = [
        "part_number",
        "manufacturer",
        "manufacturer_part_number",
        "description",
        "quantity",
        "references",
        "source_rows",
    ]
    with path.open("w", encoding="utf-8", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def write_html_report(path: Path, report: dict[str, Any]) -> None:
    """Write a portable HTML summary without external assets or scripts."""
    summary = report["summary"]
    issue_rows = "".join(
        "<tr>"
        f"<td><span class='badge {html.escape(issue['severity'])}'>{html.escape(issue['severity'])}</span></td>"
        f"<td>{html.escape(issue['code'])}</td>"
        f"<td>{html.escape(issue['message'])}</td>"
        f"<td>{html.escape(', '.join(map(str, issue['rows'])) or '-')}</td>"
        "</tr>"
        for issue in report["issues"]
    ) or "<tr><td colspan='4'>No issues found.</td></tr>"
    document = f"""<!doctype html>
<html lang="en"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Electronic BOM Analysis Report</title><style>
:root{{--ink:#182230;--muted:#667085;--line:#e4e7ec;--card:#fff;--bg:#f7f8fa;--accent:#175cd3}}
*{{box-sizing:border-box}}body{{margin:0;background:var(--bg);color:var(--ink);font:15px/1.5 system-ui,sans-serif}}
main{{max-width:1050px;margin:40px auto;padding:0 20px}}h1{{margin-bottom:4px}}.subtitle{{color:var(--muted)}}
.metrics{{display:grid;grid-template-columns:repeat(auto-fit,minmax(150px,1fr));gap:14px;margin:28px 0}}
.metric{{background:var(--card);border:1px solid var(--line);border-radius:12px;padding:18px}}.metric strong{{display:block;font-size:28px}}
table{{width:100%;border-collapse:collapse;background:var(--card);border:1px solid var(--line)}}th,td{{padding:12px;text-align:left;border-bottom:1px solid var(--line)}}th{{background:#f2f4f7}}
.badge{{display:inline-block;border-radius:999px;padding:2px 9px;font-weight:700}}.error{{background:#fee4e2;color:#b42318}}.warning{{background:#fef0c7;color:#93370d}}
footer{{margin-top:20px;color:var(--muted);font-size:13px}}</style></head><body><main>
<h1>Electronic BOM Analysis Report</h1><div class="subtitle">Generated locally. No BOM data was transmitted to an external service.</div>
<section class="metrics"><div class="metric"><span>Rows read</span><strong>{summary['rows_read']}</strong></div>
<div class="metric"><span>Unique components</span><strong>{summary['unique_components']}</strong></div>
<div class="metric"><span>Total quantity</span><strong>{summary['total_quantity']}</strong></div>
<div class="metric"><span>Errors / warnings</span><strong>{summary['errors']} / {summary['warnings']}</strong></div></section>
<h2>Validation issues</h2><table><thead><tr><th>Severity</th><th>Code</th><th>Message</th><th>Source rows</th></tr></thead><tbody>{issue_rows}</tbody></table>
<footer>Review all flagged items against the schematic, datasheets, and approved procurement requirements.</footer>
</main></body></html>"""
    path.write_text(document, encoding="utf-8")
